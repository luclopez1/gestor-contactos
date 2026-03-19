# ============================================================
#  GESTOR DE CONTACTOS WEB
#  Abre una página en tu navegador para añadir contactos
#  y los guarda automáticamente en contactos.xlsx
#
#  REQUIERE:
#  pip install flask openpyxl
# ============================================================

from flask import Flask, request, redirect, url_for
import openpyxl
import os

app = Flask(__name__)

ARCHIVO = "contactos.xlsx"

# ----------------------------------------------------------
# Crear el Excel si no existe
# ----------------------------------------------------------
def inicializar_excel():
    if not os.path.exists(ARCHIVO):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contactos"
        ws.append(["nombre", "empresa", "email"])
        wb.save(ARCHIVO)

# ----------------------------------------------------------
# Leer todos los contactos del Excel
# ----------------------------------------------------------
def leer_contactos():
    wb = openpyxl.load_workbook(ARCHIVO)
    ws = wb.active
    contactos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0]:
            contactos.append({"nombre": fila[0], "empresa": fila[1], "email": fila[2]})
    return contactos

# ----------------------------------------------------------
# Guardar un contacto nuevo en el Excel
# ----------------------------------------------------------
def guardar_contacto(nombre, empresa, email):
    wb = openpyxl.load_workbook(ARCHIVO)
    ws = wb.active
    ws.append([nombre, empresa, email])
    wb.save(ARCHIVO)

# ----------------------------------------------------------
# Eliminar un contacto por número de fila
# ----------------------------------------------------------
def eliminar_contacto(indice):
    wb = openpyxl.load_workbook(ARCHIVO)
    ws = wb.active
    ws.delete_rows(indice + 2)  # +2 porque la fila 1 es cabecera
    wb.save(ARCHIVO)

# ----------------------------------------------------------
# PÁGINA PRINCIPAL
# ----------------------------------------------------------
@app.route("/")
def index():
    contactos = leer_contactos()
    total = len(contactos)

    filas = ""
    for i, c in enumerate(contactos):
        filas += f"""
        <tr>
            <td>{i + 1}</td>
            <td>{c['nombre']}</td>
            <td>{c['empresa'] or '-'}</td>
            <td>{c['email']}</td>
            <td>
                <a href="/eliminar/{i}" 
                   onclick="return confirm('Eliminar a {c['nombre']}?')"
                   style="color:#e74c3c;text-decoration:none;font-weight:bold">
                   X Eliminar
                </a>
            </td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Gestor de Contactos</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ font-family: Arial, sans-serif; background: #f5f7fa; color: #333; }}

            header {{
                background: #2F75B6;
                color: white;
                padding: 20px 40px;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }}
            header h1 {{ font-size: 22px; }}
            header span {{ font-size: 14px; opacity: 0.8; }}

            .container {{ max-width: 900px; margin: 40px auto; padding: 0 20px; }}

            .card {{
                background: white;
                border-radius: 10px;
                padding: 30px;
                margin-bottom: 30px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            }}

            .card h2 {{
                font-size: 18px;
                color: #2F75B6;
                margin-bottom: 20px;
                padding-bottom: 10px;
                border-bottom: 2px solid #e8f0fe;
            }}

            .form-row {{
                display: flex;
                gap: 16px;
                flex-wrap: wrap;
            }}

            .form-group {{
                flex: 1;
                min-width: 200px;
                display: flex;
                flex-direction: column;
                gap: 6px;
            }}

            label {{ font-size: 13px; font-weight: bold; color: #555; }}

            input {{
                padding: 10px 14px;
                border: 1px solid #ddd;
                border-radius: 6px;
                font-size: 14px;
                outline: none;
                transition: border 0.2s;
            }}

            input:focus {{ border-color: #2F75B6; }}

            button {{
                margin-top: 20px;
                background: #2F75B6;
                color: white;
                border: none;
                padding: 12px 30px;
                border-radius: 6px;
                font-size: 15px;
                cursor: pointer;
                font-weight: bold;
            }}

            button:hover {{ background: #1F5A96; }}

            .badge {{
                background: #e8f0fe;
                color: #2F75B6;
                padding: 4px 12px;
                border-radius: 20px;
                font-size: 13px;
                font-weight: bold;
                margin-left: 10px;
            }}

            table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 14px;
            }}

            th {{
                background: #2F75B6;
                color: white;
                padding: 12px 16px;
                text-align: left;
                font-weight: bold;
            }}

            td {{
                padding: 11px 16px;
                border-bottom: 1px solid #f0f0f0;
            }}

            tr:nth-child(even) td {{ background: #f9fbff; }}
            tr:hover td {{ background: #e8f0fe; }}

            .empty {{
                text-align: center;
                color: #aaa;
                padding: 40px;
                font-size: 15px;
            }}

            .success {{
                background: #e8f8f0;
                border: 1px solid #27ae60;
                color: #1e7b34;
                padding: 12px 20px;
                border-radius: 6px;
                margin-bottom: 20px;
                font-size: 14px;
            }}

            .download-btn {{
                background: #27ae60;
                color: white;
                padding: 10px 20px;
                border-radius: 6px;
                text-decoration: none;
                font-size: 14px;
                font-weight: bold;
                display: inline-block;
            }}

            .download-btn:hover {{ background: #1e7b34; }}
        </style>
    </head>
    <body>
        <header>
            <h1>Gestor de Contactos</h1>
            <span>Los datos se guardan en contactos.xlsx</span>
        </header>

        <div class="container">

            <!-- FORMULARIO -->
            <div class="card">
                <h2>Añadir nuevo contacto</h2>
                <form method="POST" action="/añadir">
                    <div class="form-row">
                        <div class="form-group">
                            <label>Nombre *</label>
                            <input type="text" name="nombre" placeholder="Ej: Carlos García" required>
                        </div>
                        <div class="form-group">
                            <label>Empresa</label>
                            <input type="text" name="empresa" placeholder="Ej: Tech Solutions SL">
                        </div>
                        <div class="form-group">
                            <label>Email *</label>
                            <input type="email" name="email" placeholder="Ej: carlos@empresa.com" required>
                        </div>
                    </div>
                    <button type="submit">+ Añadir contacto</button>
                </form>
            </div>

            <!-- LISTA DE CONTACTOS -->
            <div class="card">
                <h2>
                    Contactos guardados
                    <span class="badge">{total}</span>
                    {"&nbsp;&nbsp;<a href='/descargar' class='download-btn'>⬇ Descargar Excel</a>" if total > 0 else ""}
                </h2>

                {"<table><thead><tr><th>#</th><th>Nombre</th><th>Empresa</th><th>Email</th><th>Acción</th></tr></thead><tbody>" + filas + "</tbody></table>" if total > 0 else '<p class="empty">No hay contactos todavía. Añade el primero arriba.</p>'}
            </div>

        </div>
    </body>
    </html>
    """
    return html

# ----------------------------------------------------------
# AÑADIR CONTACTO
# ----------------------------------------------------------
@app.route("/añadir", methods=["POST"])
def añadir():
    nombre = request.form.get("nombre", "").strip()
    empresa = request.form.get("empresa", "").strip()
    email = request.form.get("email", "").strip()

    if nombre and email:
        guardar_contacto(nombre, empresa, email)

    return redirect(url_for("index"))

# ----------------------------------------------------------
# ELIMINAR CONTACTO
# ----------------------------------------------------------
@app.route("/eliminar/<int:indice>")
def eliminar(indice):
    eliminar_contacto(indice)
    return redirect(url_for("index"))

# ----------------------------------------------------------
# DESCARGAR EXCEL
# ----------------------------------------------------------
@app.route("/descargar")
def descargar():
    from flask import send_file
    return send_file(ARCHIVO, as_attachment=True)

# ----------------------------------------------------------
# ARRANCAR LA APP
# ----------------------------------------------------------
if __name__ == "__main__":
    inicializar_excel()
    print("\n✅ Gestor de contactos iniciado.")
    print("   Abre tu navegador y ve a: http://localhost:5000")
    print("   Pulsa Ctrl+C para cerrar.\n")
    app.run(debug=False)
